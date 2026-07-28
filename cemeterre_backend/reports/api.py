# projet_cimetiere/cemeterre_backend/reports/api.py

from ninja import Router
from ninja_jwt.authentication import JWTAuth
from django.http import HttpResponse
from django.db.models import Sum
from datetime import datetime, timedelta

from cemetery.models import Grave, Section
from reservations.models import Reservation
from finance.models import Facture

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill

from .schemas import DashboardStatsSchema
from core.permissions import require_role

router = Router(auth=JWTAuth(), tags=["Reports"])


# =========================
# DASHBOARD STATISTIQUES (Conforme CDC)
# =========================
@router.get("/dashboard", response=DashboardStatsSchema)
@require_role("admin", "secretariat")
def dashboard_stats(request, cemetery_id: int = None):

    # 1. STATISTIQUES GLOBALES DES CAVEAUX
    graves = Grave.objects.all()
    if cemetery_id:
        graves = graves.filter(section__cemetery_id=cemetery_id)

    total = graves.count()
    free = graves.filter(status="available").count()
    reserved = graves.filter(status="reserved").count()
    occupied = graves.filter(status="occupied").count()
    unavailable = graves.filter(status="non_exploitable").count()

    occupation_rate = (occupied / total * 100) if total else 0
    saturation_rate = ((occupied + reserved) / total * 100) if total else 0

    # 2. TAUX D'OCCUPATION PAR SECTION (Conforme CDC)
    sections = Section.objects.all()
    if cemetery_id:
        sections = sections.filter(cemetery_id=cemetery_id)
    
    sections_stats = []
    for section in sections:
        section_graves = Grave.objects.filter(section=section).count()
        section_occupied = Grave.objects.filter(
            section=section, 
            status__in=['occupied', 'reserved']
        ).count()
        section_rate = round((section_occupied / section_graves * 100), 1) if section_graves > 0 else 0
        
        sections_stats.append({
            "id": section.id,
            "name": section.name,
            "total": section_graves,
            "occupied": section_occupied,
            "available": section_graves - section_occupied,
            "rate": section_rate
        })

    # 3. REVENUS FINANCIERS (Conforme CDC)
    # ✅ CORRECTION 1 : montant_paye → montant_total
    total_revenue = Facture.objects.filter(
        statut="payee"
    ).aggregate(total=Sum('montant_total'))['total'] or 0
    
    # Revenus des 6 derniers mois
    six_months_ago = datetime.now() - timedelta(days=180)
    monthly_revenue = []
    
    for i in range(6):
        month_start = six_months_ago + timedelta(days=30*i)
        month_end = month_start + timedelta(days=30)
        
        # ✅ CORRECTION 2 : date_creation → created_at ET montant_paye → montant_total
        month_total = Facture.objects.filter(
            statut="payee",
            created_at__gte=month_start,
            created_at__lt=month_end
        ).aggregate(total=Sum('montant_total'))['total'] or 0
        
        monthly_revenue.append({
            "month": month_start.strftime("%b %Y"),
            "amount": float(month_total)
        })

    # 4. RÉPONSE FINALE CONFORME CDC
    return {
        "total_graves": total,
        "free": free,
        "reserved": reserved,
        "occupied": occupied,
        "unavailable": unavailable,
        "occupation_rate": round(occupation_rate, 2),
        "saturation_rate": round(saturation_rate, 2),
        "sections": sections_stats,
        "total_revenue": float(total_revenue),
        "monthly_revenue": monthly_revenue,
    }


# =========================
# EXPORT GRAVES CSV
# =========================
@router.get("/export/graves/csv")
@require_role("admin", "secretariat")
def export_graves_csv(request):

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="graves.csv"'

    writer = csv.writer(response)

    writer.writerow([
        "ID",
        "Code",
        "Statut",
        "Section",
        "Cimetière",
        "Date création"
    ])

    graves = Grave.objects.select_related(
        "section",
        "section__cemetery"
    ).all()

    for g in graves:
        writer.writerow([
            g.id,
            g.code,
            g.get_status_display(),
            g.section.name,
            g.section.cemetery.name,
            g.created_at.strftime("%d/%m/%Y"),
        ])

    return response


# =========================
# EXPORT GRAVES EXCEL
# =========================
@router.get("/export/graves/excel")
@require_role("admin", "secretariat")
def export_graves_excel(request):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Graves"

    headers = [
        "ID",
        "Code",
        "Statut",
        "Section",
        "Cimetière",
        "Date création"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            start_color="1565C0",
            end_color="1565C0",
            fill_type="solid"
        )

    graves = Grave.objects.select_related(
        "section",
        "section__cemetery"
    ).all()

    for g in graves:
        ws.append([
            g.id,
            g.code,
            g.get_status_display(),
            g.section.name,
            g.section.cemetery.name,
            g.created_at.strftime("%d/%m/%Y"),
        ])

    for column in ws.columns:
        max_len = max(
            len(str(c.value or ""))
            for c in column
        )

        ws.column_dimensions[
            column[0].column_letter
        ].width = max_len + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="graves.xlsx"'

    wb.save(response)

    return response


# =========================
# EXPORT RESERVATIONS CSV
# =========================
@router.get("/export/reservations/csv")
@require_role("admin", "secretariat")
def export_reservations_csv(request):

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="reservations.csv"'

    writer = csv.writer(response)

    writer.writerow([
        "ID",
        "Client",
        "Email",
        "Grave",
        "Statut",
        "Défunt",
        "Date"
    ])

    reservations = Reservation.objects.select_related(
        "user",
        "grave"
    ).all()

    for r in reservations:
        writer.writerow([
            r.id,
            r.user.username,
            r.user.email,
            r.grave.code,
            r.get_status_display(),
            f"{r.deceased_last_name or ''} {r.deceased_first_name or ''}".strip(),
            r.reservation_date.strftime("%d/%m/%Y"),
        ])

    return response


# =========================
# EXPORT RESERVATIONS EXCEL
# =========================
@router.get("/export/reservations/excel")
@require_role("admin", "secretariat")
def export_reservations_excel(request):

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Reservations"

    headers = [
        "ID",
        "Client",
        "Email",
        "Grave",
        "Statut",
        "Défunt",
        "Date"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            start_color="1565C0",
            end_color="1565C0",
            fill_type="solid"
        )

    reservations = Reservation.objects.select_related(
        "user",
        "grave"
    ).all()

    for r in reservations:
        ws.append([
            r.id,
            r.user.username,
            r.user.email,
            r.grave.code,
            r.get_status_display(),
            f"{r.deceased_last_name or ''} {r.deceased_first_name or ''}".strip(),
            r.reservation_date.strftime("%d/%m/%Y"),
        ])

    for column in ws.columns:

        max_len = max(
            len(str(c.value or ""))
            for c in column
        )

        ws.column_dimensions[
            column[0].column_letter
        ].width = max_len + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="reservations.xlsx"'

    wb.save(response)

    return response


# =========================
# EXPORT FACTURES EXCEL
# =========================
@router.get("/export/factures/excel")
@require_role("admin", "secretariat")
def export_factures_excel(request):

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Factures"

    headers = [
        "ID",
        "Numéro",
        "Client",
        "Montant Total",
        "Montant Payé",
        "Restant",
        "Statut",
        "Date Création",
        "Date Échéance"
    ]

    ws.append(headers)

    # Style de l'en-tête
    for cell in ws[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            start_color="1565C0",
            end_color="1565C0",
            fill_type="solid"
        )

    # Récupération des factures avec les relations
    factures = Facture.objects.select_related(
        "client",
        "reservation"
    ).all()

    for f in factures:
        ws.append([
            f.id,
            f.numero,
            f.client.username if f.client else "N/A",
            float(f.montant_total or 0),
            float(getattr(f, 'montant_paye', 0) or 0),
            float(getattr(f, 'montant_restant', 0) or 0),
            f.statut,
            f.created_at.strftime("%d/%m/%Y"),
            f.date_echeance.strftime("%d/%m/%Y") if f.date_echeance else "",
        ])

    # Ajustement automatique de la largeur des colonnes
    for column in ws.columns:
        max_len = max(
            len(str(c.value or ""))
            for c in column
        )

        ws.column_dimensions[
            column[0].column_letter
        ].width = max_len + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="factures.xlsx"'

    wb.save(response)

    return response

# ==============================================================================
# EXPORT CONCESSIONS CSV
# ==============================================================================
@router.get("/export/concessions/csv")
@require_role("admin", "secretariat")
def export_concessions_csv(request):
    """Export CSV des concessions"""
    from cemetery.models import Concession
    
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="concessions.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "ID", "Contrat", "Caveau", "Section", "Client", "Email",
        "Type", "Montant", "Date début", "Date fin", "Durée (ans)",
        "Statut", "Jours restants", "Renouvellements", "Date création"
    ])

    concessions = Concession.objects.select_related("grave", "user", "grave__section").all()

    for c in concessions:
        writer.writerow([
            c.id,
            f"CONC-{c.id:06d}",
            c.grave.code,
            c.grave.section.name if c.grave.section else "",
            c.user.username,
            c.user.email,
            c.get_type_concession_display(),
            float(c.montant),
            c.date_debut.strftime("%d/%m/%Y"),
            c.date_fin.strftime("%d/%m/%Y") if c.date_fin else "Perpétuelle",
            c.duree_annees or "Perpétuelle",
            c.get_status_display(),
            c.days_remaining if c.days_remaining is not None else "N/A",
            c.renewed_count,
            c.created_at.strftime("%d/%m/%Y"),
        ])

    return response


# ==============================================================================
# EXPORT CONCESSIONS EXCEL
# ==============================================================================
@router.get("/export/concessions/excel")
@require_role("admin", "secretariat")
def export_concessions_excel(request):
    """Export Excel des concessions"""
    from cemetery.models import Concession
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Concessions"

    headers = [
        "ID", "Contrat", "Caveau", "Section", "Client", "Email",
        "Type", "Montant (FCFA)", "Date début", "Date fin", "Durée (ans)",
        "Statut", "Jours restants", "Renouvellements", "Date création"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="496042", end_color="496042", fill_type="solid")

    concessions = Concession.objects.select_related("grave", "user", "grave__section").all()

    for c in concessions:
        ws.append([
            c.id,
            f"CONC-{c.id:06d}",
            c.grave.code,
            c.grave.section.name if c.grave.section else "",
            c.user.username,
            c.user.email,
            c.get_type_concession_display(),
            float(c.montant),
            c.date_debut.strftime("%d/%m/%Y"),
            c.date_fin.strftime("%d/%m/%Y") if c.date_fin else "Perpétuelle",
            c.duree_annees or "Perpétuelle",
            c.get_status_display(),
            c.days_remaining if c.days_remaining is not None else "N/A",
            c.renewed_count,
            c.created_at.strftime("%d/%m/%Y"),
        ])

    for column in ws.columns:
        max_len = max(len(str(c.value or "")) for c in column)
        ws.column_dimensions[column[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="concessions.xlsx"'
    wb.save(response)
    return response


# ==============================================================================
# STATS CONCESSIONS
# ==============================================================================
@router.get("/concessions/stats")
@require_role("admin")
def concessions_stats_report(request):
    """Statistiques concessions pour le reporting"""
    from cemetery.models import Concession
    from django.db.models import Sum
    from datetime import timedelta
    
    today = timezone.now().date()
    
    total = Concession.objects.count()
    actives = Concession.objects.filter(status="active").count()
    expired = Concession.objects.filter(status="expired").count()
    resiliees = Concession.objects.filter(status="resiliee").count()
    temporaires = Concession.objects.filter(type_concession="temporaire").count()
    perpetuelles = Concession.objects.filter(type_concession="perpetuelle").count()
    
    expiring_15 = Concession.objects.filter(
        status="active",
        type_concession="temporaire",
        date_fin__lte=today + timedelta(days=15),
        date_fin__gte=today
    ).count()
    
    revenus = Concession.objects.filter(is_paid=True).aggregate(total=Sum('montant'))['total'] or 0
    
    return {
        "total": total,
        "actives": actives,
        "expired": expired,
        "resiliees": resiliees,
        "temporaires": temporaires,
        "perpetuelles": perpetuelles,
        "expiring_in_15_days": expiring_15,
        "revenus_total": float(revenus),
        "taux_renouvellement": round(
            (Concession.objects.filter(renewed_count__gt=0).count() / total * 100) if total > 0 else 0, 
            2
        )
    }

# ==============================================================================
# GESTION DES LOGS D'AUDIT (AUDIT TRAIL) - CDC Section 4 & 7
# ==============================================================================
from cemetery.models import AuditLog
from django.db.models import Q
from datetime import datetime, timedelta

@router.get("/audit-logs")
@require_role("admin", "secretariat")
def list_audit_logs(request, 
                    action: str = None, 
                    model_name: str = None, 
                    days: int = 30):
    """
    Liste les logs d'audit avec filtres.
    Par défaut, affiche les 30 derniers jours.
    """
    threshold_date = datetime.now() - timedelta(days=days)
    
    queryset = AuditLog.objects.filter(timestamp__gte=threshold_date).select_related("user")
    
    if action:
        queryset = queryset.filter(action=action)
    if model_name:
        queryset = queryset.filter(model_name__icontains=model_name)
        
    # On retourne les 100 derniers logs triés par date décroissante
    logs = queryset.order_by("-timestamp")[:100]
    
    return [
        {
            "id": log.id,
            "timestamp": log.timestamp.isoformat(),
            "user": log.user.username if log.user else "Système",
            "action": log.action,
            "model": log.model_name,
            "object_id": log.object_id,
            "details": log.details
        }
        for log in logs
    ]


@router.get("/audit-logs/export")
@require_role("admin", "secretariat")
def export_audit_logs_csv(request, days: int = 365):
    """
    Exporte les logs au format CSV pour archivage légal.
    Par défaut, exporte l'année complète (365 jours).
    """
    threshold_date = datetime.now() - timedelta(days=days)
    logs = AuditLog.objects.filter(timestamp__gte=threshold_date).select_related("user").order_by("-timestamp")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="audit_logs_{datetime.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)
    # En-têtes CSV
    writer.writerow([
        "ID", "Date & Heure", "Utilisateur", "Action", 
        "Modèle Cible", "ID Objet", "Détails / Justification"
    ])

    for log in logs:
        writer.writerow([
            log.id,
            log.timestamp.strftime("%d/%m/%Y %H:%M:%S"),
            log.user.username if log.user else "Système",
            log.get_action_display(), # Affiche le nom lisible (ex: "Création")
            log.model_name,
            log.object_id or "",
            log.details or ""
        ])

    return response


@router.post("/audit-logs/purge")
@require_role("admin") # SEUL L'ADMIN PEUT PURGER
def purge_old_audit_logs(request, days: int = 365):
    """
    Supprime les logs plus anciens que X jours.
    Sécurité : days ne peut pas être inférieur à 30 jours pour éviter les suppressions accidentelles.
    """
    if days < 30:
        return {"error": "La période de rétention ne peut pas être inférieure à 30 jours (conformité légale)."}

    threshold_date = datetime.now() - timedelta(days=days)
    
    # Compte les éléments à supprimer avant de les effacer
    count_to_delete = AuditLog.objects.filter(timestamp__lt=threshold_date).count()
    
    if count_to_delete == 0:
        return {"message": f"Aucun log à supprimer pour la période de plus de {days} jours."}

    # Suppression effective
    AuditLog.objects.filter(timestamp__lt=threshold_date).delete()

    # On logue cette action de purge elle-même ! (Méta-audit)
    AuditLog.objects.create(
        user=request.auth,
        action="delete",
        model_name="AuditLog",
        details=f"Purge automatique de {count_to_delete} logs de plus de {days} jours."
    )

    return {
        "message": f"Purge réussie. {count_to_delete} anciens logs ont été archivés/supprimés.",
        "deleted_count": count_to_delete
    }